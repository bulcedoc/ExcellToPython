from django.shortcuts import render
from django.shortcuts import HttpResponse
from django.shortcuts import redirect
cid = []
cname = []
cadd = []
cph = []
cemail = []
cdn = []
cdh =[]
shcustnid = []
shcustname = []
shcustnum = []
shemail = []
shStocknid =[]
shStockdesp = []
shstocktyp =[]
shpayid = []
shpaydate = []
shpaydesp = []
stsalesnid = []
stSdesp = []
stSamount = []
stsalesname =[]
stprodnum = []
stprodesp=[]
stprotyp=[]
stprodit=[]
stpayid=[]
stpaydate=[]
stpaydesp=[]
studnid = []
studname = []
studadd = []
stuEmail=[]
studphnum=[]
studdept =[]
Studdob=[]
# Create your views here.
def home(request):
    return render(request,'test.html')
def col(request):
    if request.method == 'POST':
        cid.append(request.POST.get('clgnid'))
        cname.append(request.POST.get('clgname'))
        cadd.append(request.POST.get('clgadd'))
        cph.append(request.POST.get('clgphnum'))
        cemail.append(request.POST.get('clgEmail'))
        cdn.append(request.POST.get('depname'))
        cdh.append(request.POST.get('depthead'))
        return render(request,'college.html')   
    return render(request,'college.html')
def col_c(request):
 import openpyxl
 from openpyxl import Workbook
 res = HttpResponse(content_type= 'application/ms-excel')
 res['Content-Disposition'] = 'attachment; filename= "college.xls"'
 wb = Workbook() 
 sheet = wb.active 
 sheet['A1'] = " College ID " 
 sheet['B1'] = " College name " 
 sheet['C1'] = " College Address " 
 sheet['D1'] = " College Contact " 
 sheet['E1'] =" College Email " 
 sheet['F1'] =" Department Name " 
 sheet['G1'] =" Department Head "
 data =[]
 i= 0
 while i<(len(cid)):
  d = (cid[i],cname[i],cadd[i],cph[i],cemail[i],cdn[i],cdh[i])
  data.append(d)
  print(d)
  i = i+1
 for i in data: 
    sheet.append(i) 
 wb.save(res) 
 return res


def shop(request):
    if request.method == 'POST':
        shcustnid.append(request.POST.get('custnid'))
        shcustname.append(request.POST.get('custname'))
        shcustnum.append(request.POST.get('custnum'))
        shemail.append(request.POST.get('email'))
        shStocknid.append(request.POST.get('Stocknid'))
        shStockdesp.append(request.POST.get('Stockdesp'))
        shstocktyp.append(request.POST.get('stocktyp'))
        shpayid.append(request.POST.get('payid'))
        shpaydate.append(request.POST.get('paydate'))
        shpaydesp.append(request.POST.get('paydesp'))
        return render(request,'shop.html')
    return render(request,'shop.html')
def shop_c(request):
 import openpyxl
 from openpyxl import Workbook
 res = HttpResponse(content_type= 'application/ms-excel')
 res['Content-Disposition'] = 'attachment; filename= "shop.xls"'
 wb = Workbook() 
 sheet = wb.active 
 sheet['A1'] = " CustID " 
 sheet['B1'] = " CustName " 
 sheet['C1'] = " Cust Phone number " 
 sheet['D1'] = " Cust Email " 
 sheet['E1'] =" StockID " 
 sheet['F1'] =" StockDescription " 
 sheet['G1'] =" StockType "
 sheet['H1'] = " Payment ID " 
 sheet['I1'] =" Payment Date " 
 sheet['J1'] =" Payment Description " 
 data =[]
 i= 0
 while i<(len(shcustnid)):
  d = (shcustnid[i],shcustname[i],shcustnum[i],shemail[i],shStocknid[i],shStockdesp[i],shstocktyp[i],shpayid[i],shpaydate[i],shpaydesp[i])
  data.append(d)
  i = i+1
 for i in data: 
    sheet.append(i) 
 wb.save(res) 

 return res

def store(request):
    if request.method == 'POST':
        stsalesnid.append(request.POST.get('salesnid'))
        stsalesname.append(request.POST.get('salesname'))
        stSdesp.append(request.POST.get('Sdesp'))
        stSamount.append(request.POST.get('Samount'))
        stprodnum.append(request.POST.get('prodnum'))
        stprodesp.append(request.POST.get('prodesp'))
        stprotyp.append(request.POST.get('protyp'))
        stprodit.append(request.POST.get('prodit'))
        stpayid.append(request.POST.get('payid'))
        stpaydate.append(request.POST.get('paydate'))
        stpaydesp.append(request.POST.get('paydesp'))
        
        return render(request,'store.html')
    return render(request,'store.html')
def store_c(request):
 import openpyxl
 from openpyxl import Workbook
 res = HttpResponse(content_type= 'application/ms-excel')
 res['Content-Disposition'] = 'attachment; filename= "store.xls"'
 wb = Workbook() 
 sheet = wb.active 
 sheet['A1'] = " Sales ID " 
 sheet['B1'] = " Sales Type " 
 sheet['C1'] = " Sales Description " 
 sheet['D1'] = " Sales Amount " 
 sheet['E1'] =" Product Number " 
 sheet['F1'] =" Product Description " 
 sheet['G1'] =" Product Type "
 sheet['H1'] = " Product Items " 
 sheet['I1'] =" Payment ID " 
 sheet['J1'] =" Payment Date " 
 sheet['K1'] =" Payment Description "
 data =[]
 i= 0
 while i<(len(stsalesnid)):
  d = (stsalesnid[i],stSdesp[i],stSamount[i],stsalesname[i],stprodnum[i],stprodesp[i],stprotyp[i],stprodit[i],stpayid[i],stpaydate[i],stpaydesp[i])
  data.append(d)
  i = i+1
 for i in data: 
    sheet.append(i) 
 wb.save(res) 

 return res

def student(request):
    if request.method == 'POST':
        studnid.append(request.POST.get('studnid'))
        studname.append(request.POST.get('studname'))
        studadd.append(request.POST.get('studadd'))
        Studdob.append(request.POST.get('Studdob'))
        studdept.append(request.POST.get('studdept'))
        studphnum.append(request.POST.get('studphnum'))
        stuEmail.append(request.POST.get('Email'))
        return render(request,'stu.html')
    return render(request,'stu.html')
def stu_c(request):
 import openpyxl
 from openpyxl import Workbook
 res = HttpResponse(content_type= 'application/ms-excel')
 res['Content-Disposition'] = 'attachment; filename= "student.xls"'
 wb = Workbook() 
 sheet = wb.active 
 sheet['A1'] = " Student ID " 
 sheet['B1'] = " Student name " 
 sheet['C1'] = " Student Address " 
 sheet['D1'] = " Student DOB " 
 sheet['E1'] =" Student Department " 
 sheet['F1'] =" Student Contact " 
 sheet['G1'] =" Student Email "
 data =[]
 i= 0
 while i<(len(studnid)):
  d = (studnid[i],studname[i],studadd[i],Studdob[i],studdept[i],studphnum[i],stuEmail[i])
  data.append(d)
  i = i+1
 for i in data: 
    sheet.append(i) 
 wb.save(res) 

 return res

def checkout(request):
    return render(request,'form.html')

def log(request):
    from app.models import User
    if request.method == 'POST':
        name = request.POST.get('name')
        pass1 = request.POST.get('pass')
        data = User.objects.filter(username=name,password=pass1)
        if (data):
            f = redirect('/')
            return f
        else:
              return render(request, 'login1.html')
    return render(request, 'login.html')

def sign(request):
    from app.models import User
    if request.method == 'POST':
        name = request.POST.get('username')
        pass1 = request.POST.get('createpass')
        pass2 = request.POST.get('verifiedpass')
        data = User.objects.filter(username=name)
        if (pass1!=pass2):
            return render(request, 'sign1.html')
        elif(data):
            print(data)
            return render(request, 'sign1.html')
        else:
             b = User(username=name, password=pass1)
             b.save()
             f = redirect('/login')
             return f


    return render(request, 'sign.html')
